import random
import string
import re
from datetime import timedelta, datetime
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from rest_framework import permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from .models import StudentProfile, Course, LessonProgress, Certificate
from .serializers import UserSerializer, StudentSerializer, CourseSerializer, CertificateSerializer
from .utils import generate_pdf_and_qr

User = get_user_model()

# Helper to normalize Uzbek characters to clean English letters
def normalize_uzbek_name(name):
    name = name.lower().strip()
    name = name.replace("o'", "o").replace("g'", "g").replace("o`", "o").replace("g`", "g")
    name = name.replace("sh", "sh").replace("ch", "ch")
    name = re.sub(r'[^a-zA-Z0-9]', '_', name)
    return name

# Helper to generate unique username matching ismi_familiyasi
def generate_username(first_name, last_name):
    first = normalize_uzbek_name(first_name)
    last = normalize_uzbek_name(last_name)
    base_username = f"{first}_{last}"
    
    # Trim excessive underscores
    base_username = re.sub(r'_+', '_', base_username).strip('_')
    
    if not base_username:
        base_username = "user"
        
    username = base_username
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f"{base_username}_{counter}"
        counter += 1
    return username

# Helper to check if user has admin permission
def is_admin_user(user):
    return user.role == User.Role.ADMIN or user.is_staff or user.is_superuser

# Helper to check if user has teacher permission
def is_teacher_user(user):
    return user.role == User.Role.TEACHER or is_admin_user(user)

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def auth_me(request):
    """
    Get details of the currently authenticated user.
    """
    serializer = UserSerializer(request.user, context={'request': request})
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def change_password(request):
    """
    Change password of the logged-in user.
    Payload: {"current_password": "...", "new_password": "..."}
    """
    current_password = request.data.get('current_password')
    new_password = request.data.get('new_password')
    
    if not current_password or not new_password:
        return Response({"detail": "Amaldagi parol va yangi parol kiritilishi shart."}, status=status.HTTP_400_BAD_REQUEST)
        
    if not request.user.check_password(current_password):
        return Response({"detail": "Amaldagi parol noto'g'ri kiritildi."}, status=status.HTTP_400_BAD_REQUEST)
        
    request.user.set_password(new_password)
    request.user.save()
    return Response({"detail": "Parol muvaffaqiyatli yangilandi."})


# --- ADMIN ENDPOINTS ---

@api_view(['GET', 'POST'])
@permission_classes([permissions.IsAuthenticated])
def admin_users(request):
    """
    GET: List all teachers and students (Admin view).
    POST: Create a new Teacher user.
    """
    if not is_admin_user(request.user):
        return Response({"detail": "Ushbu sahifani ko'rish uchun ruxsatingiz yo'q (Admin talab etiladi)."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        users = User.objects.all().order_by('-id')
        serializer = UserSerializer(users, many=True, context={'request': request})
        return Response(serializer.data)

    elif request.method == 'POST':
        first_name = request.data.get('first_name')
        last_name = request.data.get('last_name')
        email = request.data.get('email', '')
        passport_series = request.data.get('passport_series', '')
        passport_number = request.data.get('passport_number', '')
        jshshir = request.data.get('jshshir', '')
        birth_date = request.data.get('birth_date') or None
        father_name = request.data.get('father_name', '')
        role = request.data.get('role', User.Role.TEACHER)
        profile_picture = request.FILES.get('profile_picture')

        if not first_name or not last_name:
            return Response({"detail": "Ism va Familiya kiritilishi shart."}, status=status.HTTP_400_BAD_REQUEST)

        if role not in [User.Role.TEACHER, User.Role.ADMIN]:
            return Response({"detail": "Noto'g'ri rol tanlandi."}, status=status.HTTP_400_BAD_REQUEST)

        username = generate_username(first_name, last_name)
        
        new_user = User.objects.create_user(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            passport_series=passport_series.upper(),
            passport_number=passport_number,
            jshshir=jshshir,
            birth_date=birth_date,
            father_name=father_name,
            password='ses2026',
            role=role,
            profile_picture=profile_picture
        )

        serializer = UserSerializer(new_user, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def admin_reset_password(request):
    """
    Reset target user's password back to 'ses2026'. (Admin only)
    Payload: {"user_id": int}
    """
    if not is_admin_user(request.user):
        return Response({"detail": "Faqat administratorlar parollarni qayta tiklay oladilar."}, status=status.HTTP_403_FORBIDDEN)

    user_id = request.data.get('user_id')
    if not user_id:
        return Response({"detail": "Foydalanuvchi ID kiritilishi shart."}, status=status.HTTP_400_BAD_REQUEST)

    target_user = get_object_or_404(User, id=user_id)
    target_user.set_password('ses2026')
    target_user.save()

    return Response({"detail": f"Foydalanuvchi {target_user.username} paroli 'ses2026' ga qaytarildi."})


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def admin_all_certificates(request):
    """
    GET: List all certificates in the system.
    Supports Admin (all certificates) and Teacher (only their course certificates).
    Supports optional search by student name or certificate_id.
    """
    is_admin = is_admin_user(request.user)
    is_teacher = is_teacher_user(request.user)

    if not is_admin and not is_teacher:
        return Response({"detail": "Faqat administratorlar va o'qituvchilar sertifikatlarni ko'ra oladi."}, status=status.HTTP_403_FORBIDDEN)

    search = request.GET.get('search', '').strip()
    
    if is_admin:
        certs = Certificate.objects.select_related('student', 'course').order_by('-issued_at')
    else:
        certs = Certificate.objects.filter(course__teacher=request.user).select_related('student', 'course').order_by('-issued_at')

    if search:
        if is_admin:
            certs = certs.filter(
                certificate_id__icontains=search
            ) | Certificate.objects.filter(
                student__first_name__icontains=search
            ) | Certificate.objects.filter(
                student__last_name__icontains=search
            ) | Certificate.objects.filter(
                course__title__icontains=search
            )
        else:
            certs = certs.filter(
                course__teacher=request.user,
                certificate_id__icontains=search
            ) | Certificate.objects.filter(
                course__teacher=request.user,
                student__first_name__icontains=search
            ) | Certificate.objects.filter(
                course__teacher=request.user,
                student__last_name__icontains=search
            ) | Certificate.objects.filter(
                course__teacher=request.user,
                course__title__icontains=search
            )
        certs = certs.distinct().order_by('-issued_at')

    result = []
    for cert in certs:
        student = cert.student
        pic = None
        if student.profile_picture:
            pic = request.build_absolute_uri(student.profile_picture.url)
        
        # Build absolute pdf path if file exists
        pdf_path = None
        if cert.pdf_file:
            pdf_path = request.build_absolute_uri(cert.pdf_file.url)

        parts = [student.first_name, student.last_name, student.father_name]
        student_full = " ".join([p.strip() for p in parts if p and p.strip()])

        result.append({
            "certificate_id": cert.certificate_id,
            "student_id": student.id,
            "student_name": student_full or student.username,
            "student_username": student.username,
            "student_picture": pic,
            "course_id": cert.course.id,
            "course_name": cert.course.title,
            "teacher_id": cert.course.teacher.id,
            "teacher_name": cert.course.teacher.get_full_name() or cert.course.teacher.username,
            "issued_at": cert.issued_at.strftime('%d.%m.%Y') if cert.issued_at else None,
            "expires_at": cert.expires_at.strftime('%d.%m.%Y') if cert.expires_at else None,
            "is_active": cert.is_active,
            "qr_code_image": request.build_absolute_uri(cert.qr_code_image.url) if cert.qr_code_image else None,
            "pdf_file": pdf_path,
        })

    return Response(result)


# --- TEACHER ENDPOINTS ---

@api_view(['GET', 'POST'])
@permission_classes([permissions.IsAuthenticated])
def teacher_courses(request):
    """
    GET: List all courses created by the logged-in teacher (or all courses for admin).
    POST: Create a new Course.
    """
    if not is_teacher_user(request.user):
        return Response({"detail": "Darslarni boshqarish faqat o'qituvchilarga ruxsat berilgan."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        if is_admin_user(request.user):
            courses = Course.objects.all().order_by('-id')
        else:
            courses = Course.objects.filter(teacher=request.user).order_by('-id')
        serializer = CourseSerializer(courses, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        title = request.data.get('title')
        description = request.data.get('description', '')
        start_date = request.data.get('start_date')
        end_date = request.data.get('end_date')
        total_lessons = request.data.get('total_lessons', 10)

        if not title or not start_date or not end_date:
            return Response({"detail": "Kurs nomi, boshlanish va tugash sanalari majburiy."}, status=status.HTTP_400_BAD_REQUEST)

        course = Course.objects.create(
            title=title,
            description=description,
            teacher=request.user,
            start_date=start_date,
            end_date=end_date,
            total_lessons=int(total_lessons)
        )

        serializer = CourseSerializer(course)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def teacher_students(request):
    """
    GET: List all students linked to the teacher.
    Supports ?all=true to fetch all students for dropdown selection.
    """
    if not is_teacher_user(request.user):
        return Response({"detail": "Talabalarni boshqarish faqat o'qituvchilarga ruxsat berilgan."}, status=status.HTTP_403_FORBIDDEN)

    get_all = request.GET.get('all', 'false').lower() == 'true'

    if is_admin_user(request.user) or get_all:
        profiles = StudentProfile.objects.all().distinct().order_by('-id')
    else:
        from django.db.models import Q
        profiles = StudentProfile.objects.filter(
            Q(teacher=request.user) | Q(user__assigned_courses__teacher=request.user)
        ).distinct().order_by('-id')
        
    serializer = StudentSerializer(profiles, many=True, context={'request': request})
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
@transaction.atomic
def course_add_student(request, course_id):
    """
    Add a student to a specific course.
    If 'student_id' is provided, enroll that existing student.
    Otherwise, create a new student user and profile automatically.
    """
    if not is_teacher_user(request.user):
        return Response({"detail": "O'quvchilarni kursga qo'shish faqat o'qituvchilarga ruxsat etiladi."}, status=status.HTTP_403_FORBIDDEN)

    course = get_object_or_404(Course, id=course_id)
    if not is_admin_user(request.user) and course.teacher != request.user:
        return Response({"detail": "Ushbu kursni tahrirlash huquqi sizda yo'q."}, status=status.HTTP_403_FORBIDDEN)

    student_id = request.data.get('student_id')
    student_ids = request.data.get('student_ids')

    if student_ids or student_id:
        ids_to_process = student_ids if student_ids else [student_id]
        last_profile = None
        for s_id in ids_to_process:
            student_user = get_object_or_404(User, id=s_id, role=User.Role.STUDENT)
            profile = StudentProfile.objects.filter(user=student_user).first()
            # Add to course
            course.students.add(student_user)
            # Initialize LessonProgress
            LessonProgress.objects.get_or_create(
                student=student_user,
                course=course,
                defaults={'completed_lessons': []}
            )
            if profile:
                last_profile = profile
        
        if last_profile:
            serializer = StudentSerializer(last_profile, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({"detail": "O'quvchilar muvaffaqiyatli biriktirildi."})

    first_name = request.data.get('first_name')
    last_name = request.data.get('last_name')
    father_name = request.data.get('father_name', '')
    email = request.data.get('email', '')
    phone_number = request.data.get('phone_number', '')
    organization = request.data.get('organization', '')
    passport_series = request.data.get('passport_series', '')
    passport_number = request.data.get('passport_number', '')
    jshshir = request.data.get('jshshir', '')
    profile_picture = request.FILES.get('profile_picture')

    if not first_name or not last_name:
        return Response({"detail": "Ism va Familiya majburiy."}, status=status.HTTP_400_BAD_REQUEST)

    # Autogenerate username
    username = generate_username(first_name, last_name)

    # 1. Create Student User with default password 'ses2026' and profile picture
    student_user = User.objects.create_user(
        username=username,
        first_name=first_name,
        last_name=last_name,
        email=email,
        passport_series=passport_series.upper(),
        passport_number=passport_number,
        jshshir=jshshir,
        father_name=father_name,
        password='ses2026',
        role=User.Role.STUDENT,
        profile_picture=profile_picture
    )

    # 2. Create StudentProfile
    profile = StudentProfile.objects.create(
        user=student_user,
        teacher=course.teacher,
        phone_number=phone_number,
        organization=organization
    )

    # 3. Add to Course students
    course.students.add(student_user)

    # 4. Initialize LessonProgress
    LessonProgress.objects.create(
        student=student_user,
        course=course,
        completed_lessons=[]
    )

    serializer = StudentSerializer(profile, context={'request': request})
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def toggle_lesson(request, student_id):
    """
    Toggle completed status of a lesson or multiple lessons for a student in their course.
    Payload: {"course_id": int, "lesson_id": int/null, "lesson_ids": list, "completed": bool}
    """
    if not is_teacher_user(request.user):
        return Response({"detail": "Darslar holatini o'zgartirish faqat o'qituvchilarga ruxsat berilgan."}, status=status.HTTP_403_FORBIDDEN)

    course_id = request.data.get('course_id')
    lesson_id = request.data.get('lesson_id')
    lesson_ids = request.data.get('lesson_ids')
    completed = request.data.get('completed', False)

    if not course_id:
        return Response({"detail": "Course ID kiritilishi shart."}, status=status.HTTP_400_BAD_REQUEST)

    student_user = get_object_or_404(User, id=student_id, role=User.Role.STUDENT)
    course = get_object_or_404(Course, id=course_id)

    # Check permission
    if not is_admin_user(request.user) and course.teacher != request.user:
        return Response({"detail": "Ushbu kurs o'quvchilarini boshqarishga huquqingiz yo'q."}, status=status.HTTP_403_FORBIDDEN)

    progress, created = LessonProgress.objects.get_or_create(student=student_user, course=course)
    completed_list = list(progress.completed_lessons)

    # Determine target lesson ids
    target_ids = []
    if lesson_ids is not None:
        target_ids = list(lesson_ids)
    elif lesson_id is not None:
        target_ids = [lesson_id]

    completed_set = set(completed_list)
    if completed:
        completed_set.update(target_ids)
    else:
        completed_set.difference_update(target_ids)

    progress.completed_lessons = sorted(list(completed_set))
    progress.save()

    return Response({
        "student_id": student_id,
        "course_id": course_id,
        "completed_lessons": progress.completed_lessons
    })


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def generate_certificate(request, student_id):
    """
    Generate certificate for a student.
    Payload: {"course_id": int}
    """
    if not is_teacher_user(request.user):
        return Response({"detail": "Faqat o'qituvchilar sertifikat yaratishi mumkin."}, status=status.HTTP_403_FORBIDDEN)

    course_id = request.data.get('course_id')
    if not course_id:
        return Response({"detail": "Course ID majburiy."}, status=status.HTTP_400_BAD_REQUEST)

    student_user = get_object_or_404(User, id=student_id, role=User.Role.STUDENT)
    course = get_object_or_404(Course, id=course_id)

    if not is_admin_user(request.user) and course.teacher != request.user:
        return Response({"detail": "Ushbu kurs sertifikatini yaratishga huquqingiz yo'q."}, status=status.HTTP_403_FORBIDDEN)

    # Check if certificate exists
    if Certificate.objects.filter(student=student_user, course=course).exists():
        cert = Certificate.objects.filter(student=student_user, course=course).first()
        serializer = CertificateSerializer(cert, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    # ── Barcha darslarni "yakunlandi" statusiga o'tkazish ─────────────────────
    # Sertifikat yaratishdan oldin o'quvchining ushbu kurs uchun barcha darslarini
    # tugallangan deb belgilaymiz.
    total = course.total_lessons or 0
    if total > 0:
        progress, _ = LessonProgress.objects.get_or_create(student=student_user, course=course)
        all_lessons = list(range(1, total + 1))
        # Add any missing lessons to completed list
        completed_set = set(progress.completed_lessons)
        completed_set.update(all_lessons)
        progress.completed_lessons = sorted(completed_set)
        progress.save()

    # Generate certificate ID
    def generate_cert_id():
        ids = Certificate.objects.values_list('certificate_id', flat=True)
        numeric_vals = []
        for cid in ids:
            cid_upper = cid.upper()
            if cid_upper.startswith("SES-"):
                num_part = cid_upper[4:]
                if num_part.isdigit():
                    numeric_vals.append(int(num_part))
            elif cid.isdigit():
                numeric_vals.append(int(cid))
        next_val = max(numeric_vals) + 1 if numeric_vals else 1
        cert_id = f"SES-{next_val}"
        while Certificate.objects.filter(certificate_id=cert_id).exists():
            next_val += 1
            cert_id = f"SES-{next_val}"
        return cert_id

    cert_id = generate_cert_id()
    expiry_date = timezone.now() + timedelta(days=365) # Valid for 1 year

    cert = Certificate(
        certificate_id=cert_id,
        student=student_user,
        course=course,
        expires_at=expiry_date
    )

    generate_pdf_and_qr(cert)
    cert.save()

    serializer = CertificateSerializer(cert, context={'request': request})
    return Response(serializer.data, status=status.HTTP_201_CREATED)


# --- STUDENT ENDPOINTS ---

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def student_progress(request):
    """
    Get progress and certificate info of the logged-in student user.
    """
    if request.user.role != User.Role.STUDENT:
        return Response({"detail": "Ushbu sahifa faqat o'quvchilar uchun."}, status=status.HTTP_403_FORBIDDEN)

    progresses = LessonProgress.objects.filter(student=request.user)
    results = []
    
    for prog in progresses:
        cert = Certificate.objects.filter(student=request.user, course=prog.course).first()
        results.append({
            "course_id": prog.course.id,
            "course_title": prog.course.title,
            "completed_lessons": prog.completed_lessons,
            "total_lessons": prog.course.total_lessons,
            "start_date": prog.course.start_date.strftime('%d.%m.%Y'),
            "end_date": prog.course.end_date.strftime('%d.%m.%Y'),
            "has_certificate": cert is not None,
            "certificate": CertificateSerializer(cert, context={'request': request}).data if cert else None
        })

    return Response(results)


# --- PUBLIC ENDPOINTS ---

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def verify_certificate(request, certificate_id):
    """
    Verify certificate details by certificate ID.
    No auth required. Publicly accessible.
    """
    try:
        cert = Certificate.objects.get(certificate_id=certificate_id, is_active=True)
    except Certificate.DoesNotExist:
        return Response({"detail": "Sertifikat topilmadi yoki muddati yakunlangan."}, status=status.HTTP_404_NOT_FOUND)

    # Check expiration
    if cert.expires_at < timezone.now():
        cert.is_active = False
        cert.save()
        parts = [cert.student.first_name, cert.student.last_name, cert.student.father_name]
        student_full = " ".join([p.strip() for p in parts if p and p.strip()])
        return Response({
            "is_valid": False,
            "detail": "Sertifikat muddati tugagan.",
            "certificate_id": cert.certificate_id,
            "student_name": student_full or cert.student.username,
            "course_name": cert.course.title,
            "expiry_date": cert.expires_at.strftime('%d.%m.%Y')
        }, status=status.HTTP_200_OK)

    serializer = CertificateSerializer(cert, context={'request': request})
    data = serializer.data
    data['is_valid'] = True
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def verify_certificates_by_passport(request):
    """
    Verify and retrieve certificates for a student by passport series and number.
    No auth required. Publicly accessible.
    """
    passport_series = request.query_params.get('passport_series', '').strip()
    passport_number = request.query_params.get('passport_number', '').strip()

    if not passport_series or not passport_number:
        return Response({"detail": "Pasport seriyasi va raqami to'liq bo'lishi kerak."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        student = User.objects.get(
            role=User.Role.STUDENT,
            passport_series__iexact=passport_series,
            passport_number__iexact=passport_number
        )
    except User.DoesNotExist:
        return Response({"detail": "Ushbu pasport ma'lumotlariga mos o'quvchi topilmadi."}, status=status.HTTP_404_NOT_FOUND)

    # Fetch all certificates
    certs = Certificate.objects.filter(student=student)
    
    # Check and update expirations
    now = timezone.now()
    active_certs = []
    for cert in certs:
        if cert.is_active and cert.expires_at < now:
            cert.is_active = False
            cert.save()
        active_certs.append(cert)

    serializer = CertificateSerializer(active_certs, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def admin_teacher_stats(request):
    """
    Get statistics for all teachers:
    - Number of courses (darslar soni)
    - Number of registered students (o'quvchilar soni)
    - Total student-course enrollments (kursdagi o'quvchilar soni)
    - Number of issued certificates (berilgan sertifikatlar soni)
    """
    if not is_admin_user(request.user):
        return Response({"detail": "Faqat administratorlar ushbu ma'lumotlarni ko'ra oladilar."}, status=status.HTTP_403_FORBIDDEN)

    teachers = User.objects.filter(role=User.Role.TEACHER).order_by('-id')
    stats = []

    for teacher in teachers:
        courses = Course.objects.filter(teacher=teacher)
        courses_count = courses.count()
        students_count = StudentProfile.objects.filter(teacher=teacher).count()
        total_enrollments = sum(course.students.count() for course in courses)
        certificates_count = Certificate.objects.filter(course__teacher=teacher).count()

        courses_list = []
        for course in courses:
            today = timezone.now().date()
            if today < course.start_date:
                status_str = "UPCOMING"
                time_progress = 0
            elif today > course.end_date:
                status_str = "COMPLETED"
                time_progress = 100
            else:
                status_str = "ACTIVE"
                total_days = (course.end_date - course.start_date).days
                elapsed_days = (today - course.start_date).days
                time_progress = int((elapsed_days / total_days) * 100) if total_days > 0 else 100

            courses_list.append({
                "id": course.id,
                "title": course.title,
                "student_count": course.students.count(),
                "start_date": course.start_date.strftime('%d.%m.%Y'),
                "end_date": course.end_date.strftime('%d.%m.%Y'),
                "time_progress": time_progress,
                "total_lessons": course.total_lessons,
                "status": status_str
            })

        stats.append({
            "id": teacher.id,
            "username": teacher.username,
            "first_name": teacher.first_name,
            "last_name": teacher.last_name,
            "email": teacher.email,
            "courses_count": courses_count,
            "students_count": students_count,
            "total_enrollments": total_enrollments,
            "certificates_count": certificates_count,
            "courses": courses_list
        })

    return Response(stats)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def download_excel_template(request):
    """
    Generate and download a clean Excel template (.xlsx) for importing students.
    """
    if not is_teacher_user(request.user):
        return Response({"detail": "Faqat o'qituvchilarga ruxsat etiladi."}, status=status.HTTP_403_FORBIDDEN)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "O'quvchilar shabloni"
    
    # Columns
    headers = [
        "Familiya *", 
        "Ism *", 
        "Otasining ismi *", 
        "Pasport seriyasi (2 ta harf) *", 
        "Pasport raqami (7 ta raqam) *", 
        "Telefon raqami (ixtiyoriy, masalan: +998901234567)"
    ]
    
    # Add title row
    ws.append(["Tizimga o'quvchilarni qo'shish uchun shablon (* belgisi bor maydonlar majburiy)"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Add headers row
    ws.append(headers)
    ws.row_dimensions[2].height = 20
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True)
    
    for col_num in range(1, 7):
        cell = ws.cell(row=2, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Sample data
    ws.append(["Valiyev", "Ali", "Umarovich", "AA", "1234567", "+998901234567"])
    ws.append(["Karimova", "Zilola", "Anvarovna", "AB", "7654321", ""])
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    
    # Return Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=ses_oquvchilar_shablon.xlsx"
    wb.save(response)
    return response


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
@transaction.atomic
def teacher_import_excel(request, course_id):
    """
    Import students from uploaded Excel file and enroll them into the course.
    """
    if not is_teacher_user(request.user):
        return Response({"detail": "O'quvchilarni import qilish faqat o'qituvchilarga ruxsat etiladi."}, status=status.HTTP_403_FORBIDDEN)
        
    course = get_object_or_404(Course, id=course_id)
    if not is_admin_user(request.user) and course.teacher != request.user:
        return Response({"detail": "Ushbu kursni tahrirlash huquqi sizda yo'q."}, status=status.HTTP_403_FORBIDDEN)
        
    file_obj = request.FILES.get('file')
    if not file_obj:
        return Response({"detail": "Fayl yuborilmadi."}, status=status.HTTP_400_BAD_REQUEST)
        
    if not file_obj.name.endswith('.xlsx'):
        return Response({"detail": "Faqat .xlsx formatidagi Excel fayllar qabul qilinadi."}, status=status.HTTP_400_BAD_REQUEST)
        
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        
        # Read from row 3 (since row 1 is title, row 2 is headers)
        imported_count = 0
        errors = []
        
        row_num = 2
        for row in ws.iter_rows(min_row=3, values_only=True):
            row_num += 1
            # Skip empty rows
            if not row or not any(row):
                continue
                
            last_name = str(row[0]).strip() if row[0] is not None else ""
            first_name = str(row[1]).strip() if row[1] is not None else ""
            father_name = str(row[2]).strip() if row[2] is not None else ""
            passport_series = str(row[3]).strip() if row[3] is not None else ""
            passport_number = str(row[4]).strip() if row[4] is not None else ""
            phone_number = str(row[5]).strip() if row[5] is not None else ""
            
            # Validation
            if not last_name or not first_name or not father_name or not passport_series or not passport_number:
                errors.append(f"{row_num}-qatorda xatolik: Majburiy maydonlar to'ldirilmagan.")
                continue
                
            if len(passport_series) != 2:
                errors.append(f"{row_num}-qatorda xatolik: Pasport seriyasi 2 ta harf bo'lishi shart.")
                continue
                
            if len(passport_number) != 7 or not passport_number.isdigit():
                errors.append(f"{row_num}-qatorda xatolik: Pasport raqami 7 ta raqam bo'lishi shart.")
                continue
                
            # Check if student already exists by passport series & number
            student_user = User.objects.filter(
                passport_series__iexact=passport_series,
                passport_number=passport_number,
                role=User.Role.STUDENT
            ).first()
            
            if not student_user:
                # Create user
                username = generate_username(first_name, last_name)
                student_user = User.objects.create_user(
                    username=username,
                    first_name=first_name,
                    last_name=last_name,
                    father_name=father_name,
                    passport_series=passport_series.upper(),
                    passport_number=passport_number,
                    password='ses2026',
                    role=User.Role.STUDENT
                )
                # Create profile
                StudentProfile.objects.create(
                    user=student_user,
                    teacher=course.teacher,
                    phone_number=phone_number,
                    organization=""
                )
            else:
                # Ensure profile exists
                profile, _ = StudentProfile.objects.get_or_create(
                    user=student_user,
                    defaults={'teacher': course.teacher, 'phone_number': phone_number, 'organization': ""}
                )
                
            # Enroll to course
            if not course.students.filter(id=student_user.id).exists():
                course.students.add(student_user)
                # Initialize progress
                LessonProgress.objects.get_or_create(
                    student=student_user,
                    course=course,
                    defaults={'completed_lessons': []}
                )
                imported_count += 1
                
        if errors and imported_count == 0:
            return Response({"detail": "\n".join(errors)}, status=status.HTTP_400_BAD_REQUEST)
            
        return Response({
            "detail": f"{imported_count} ta o'quvchi muvaffaqiyatli import qilindi.",
            "imported_count": imported_count,
            "errors": errors
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({"detail": f"Faylni o'qishda xatolik yuz berdi: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated])
def admin_delete_user(request, user_id):
    """
    DELETE: Delete a user by ID. (Admin only)
    """
    if not is_admin_user(request.user):
        return Response({"detail": "Ushbu amalni bajarish uchun ruxsatingiz yo'q (Admin talab etiladi)."}, status=status.HTTP_403_FORBIDDEN)
        
    target_user = get_object_or_404(User, id=user_id)
    if target_user.is_superuser or target_user == request.user:
        return Response({"detail": "Tizim ma'murini yoki o'zingizni o'chira olmaysiz."}, status=status.HTTP_400_BAD_REQUEST)
        
    # Delete certificate files first
    certs = Certificate.objects.filter(student=target_user)
    for cert in certs:
        if cert.pdf_file:
            try:
                cert.pdf_file.delete(save=False)
            except:
                pass
        if cert.qr_code_image:
            try:
                cert.qr_code_image.delete(save=False)
            except:
                pass
                
    username = target_user.username
    target_user.delete()
    return Response({"detail": f"Foydalanuvchi {username} muvaffaqiyatli o'chirildi."})


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated])
def teacher_delete_student(request, student_id):
    """
    DELETE: Delete a student. (Teacher only, must be their student)
    """
    if not is_teacher_user(request.user):
        return Response({"detail": "Faqat o'qituvchilar o'quvchilarni o'chira oladi."}, status=status.HTTP_403_FORBIDDEN)
        
    student_user = get_object_or_404(User, id=student_id, role=User.Role.STUDENT)
    
    # Check if student is taught by this teacher or is in one of their courses
    is_authorized = is_admin_user(request.user)
    if not is_authorized:
        profile_exists = StudentProfile.objects.filter(user=student_user, teacher=request.user).exists()
        course_exists = Course.objects.filter(teacher=request.user, students=student_user).exists()
        is_authorized = profile_exists or course_exists
        
    if not is_authorized:
        return Response({"detail": "Ushbu o'quvchini o'chirishga ruxsatingiz yo'q."}, status=status.HTTP_403_FORBIDDEN)
        
    # Delete certificate files first
    certs = Certificate.objects.filter(student=student_user)
    for cert in certs:
        if cert.pdf_file:
            try:
                cert.pdf_file.delete(save=False)
            except:
                pass
        if cert.qr_code_image:
            try:
                cert.qr_code_image.delete(save=False)
            except:
                pass
                
    username = student_user.username
    student_user.delete()
    return Response({"detail": f"O'quvchi {username} muvaffaqiyatli o'chirildi."})


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated])
def admin_delete_certificate(request, certificate_id):
    """
    DELETE: Delete a certificate by its certificate_id. (Admin only)
    """
    if not is_admin_user(request.user):
        return Response({"detail": "Ushbu amalni bajarish uchun ruxsatingiz yo'q (Admin talab etiladi)."}, status=status.HTTP_403_FORBIDDEN)
        
    cert = get_object_or_404(Certificate, certificate_id=certificate_id)
    
    # Delete physical files from disk
    if cert.pdf_file:
        try:
            cert.pdf_file.delete(save=False)
        except:
            pass
    if cert.qr_code_image:
        try:
            cert.qr_code_image.delete(save=False)
        except:
            pass
            
    cert_id = cert.certificate_id
    cert.delete()
    return Response({"detail": f"Sertifikat {cert_id} muvaffaqiyatli o'chirildi."})


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def admin_export_certificates_excel(request):
    """
    Export certificates to an Excel sheet.
    Supports optional date filtering via query params: start_date and end_date.
    """
    if not is_admin_user(request.user):
        return Response({"detail": "Faqat administratorlar sertifikatlarni eksport qila oladilar."}, status=status.HTTP_403_FORBIDDEN)
        
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    
    certs = Certificate.objects.select_related('student', 'course', 'course__teacher').order_by('-issued_at')
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            certs = certs.filter(issued_at__date__gte=start_dt.date())
        except ValueError:
            pass
            
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            certs = certs.filter(issued_at__date__lte=end_dt.date())
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sertifikatlar"
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Columns
    headers = [
        "T/r", 
        "Sertifikat ID", 
        "F.I.Sh.", 
        "Pasport",
        "Kurs nomi", 
        "O'qituvchi", 
        "Berilgan sana",
        "Amal qilish muddati",
        "Holati"
    ]
    
    # Add title row
    ws.append(["Tizimdagi sertifikatlar ro'yxati"])
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    # emerald green style: 059669
    ws["A1"].fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Add date range if filtered
    date_info = "Barchasi"
    if start_date or end_date:
        date_info = f"Sana: {start_date or '...'} - {end_date or '...'}"
    ws.append([date_info])
    ws.merge_cells("A2:I2")
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="000000")
    ws["A2"].fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Add headers row
    ws.append(headers)
    ws.row_dimensions[3].height = 25
    header_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="0f5132")
    
    for col_num in range(1, 10):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Write data rows
    for index, cert in enumerate(certs, start=1):
        student = cert.student
        parts = [student.first_name, student.last_name, student.father_name]
        student_full = " ".join([p.strip() for p in parts if p and p.strip()]) or student.username
        
        passport = f"{student.passport_series}{student.passport_number}"
        if not passport:
            passport = "—"
            
        issued_str = cert.issued_at.strftime('%d.%m.%Y') if cert.issued_at else "—"
        expires_str = cert.expires_at.strftime('%d.%m.%Y') if cert.expires_at else "—"
        status_str = "Faol" if cert.is_active else "Yakunlangan"
        
        row_data = [
            index,
            cert.certificate_id,
            student_full,
            passport,
            cert.course.title,
            cert.course.teacher.get_full_name() or cert.course.teacher.username,
            issued_str,
            expires_str,
            status_str
        ]
        ws.append(row_data)
        
        curr_row = index + 3
        ws.row_dimensions[curr_row].height = 20
        # Alignment for cells in this row
        for col_num in range(1, 10):
            cell = ws.cell(row=curr_row, column=col_num)
            cell.font = Font(name="Calibri", size=10)
            if col_num in [1, 2, 4, 7, 8, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-adjust column widths
    from openpyxl.utils import get_column_letter
    for col_num in range(1, 10):
        col_letter = get_column_letter(col_num)
        max_len = 0
        for row_num in range(3, ws.max_row + 1):
            val = ws.cell(row=row_num, column=col_num).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"sertifikatlar_{start_date or 'barchasi'}_{end_date or 'barchasi'}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def admin_check_certificate_id(request):
    """
    GET: Check if a certificate ID already exists.
    Query param: id=ses-...
    """
    if not is_admin_user(request.user):
        return Response({"detail": "Ruxsat berilmagan."}, status=status.HTTP_403_FORBIDDEN)
        
    cert_id = request.GET.get('id', '').strip()
    if not cert_id:
        return Response({"exists": False})
        
    exists = Certificate.objects.filter(certificate_id=cert_id).exists()
    return Response({"exists": exists})


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def admin_edit_certificate_id(request, certificate_id):
    """
    POST: Edit a certificate ID (Admin only).
    Body: {"new_id": "ses-..."}
    """
    import os
    from django.conf import settings
    from .utils import generate_pdf_and_qr

    if not is_admin_user(request.user):
        return Response({"detail": "Faqat administratorlar sertifikat ID sini tahrirlashi mumkin."}, status=status.HTTP_403_FORBIDDEN)

    new_id = request.data.get('new_id', '').strip()
    if not new_id:
        return Response({"detail": "Yangi ID kiritilishi shart."}, status=status.HTTP_400_BAD_REQUEST)

    # Prefix check
    if not new_id.lower().startswith('ses-'):
        return Response({"detail": "Sertifikat ID prefiksi 'ses-' bo'lishi shart."}, status=status.HTTP_400_BAD_REQUEST)

    # Identical check
    if new_id == certificate_id:
        return Response({"detail": "Sertifikat ID o'zgartirilmadi."})

    # Duplication check
    if Certificate.objects.filter(certificate_id=new_id).exists():
        return Response({"detail": "Ushbu ID dagi sertifikat allaqachon mavjud."}, status=status.HTTP_400_BAD_REQUEST)

    cert = get_object_or_404(Certificate, certificate_id=certificate_id)

    # 1. Delete old physical files from disk
    old_pdf_path = cert.pdf_file.path if cert.pdf_file else None
    old_qr_path = cert.qr_code_image.path if cert.qr_code_image else None
    old_word_path = os.path.join(settings.MEDIA_ROOT, 'certificates', 'words', f"{certificate_id}.docx")

    # Update ID
    Certificate.objects.filter(certificate_id=certificate_id).update(certificate_id=new_id)

    # Retrieve again
    cert = Certificate.objects.get(certificate_id=new_id)

    # Disk deletion
    if old_pdf_path and os.path.exists(old_pdf_path):
        try:
            os.remove(old_pdf_path)
        except Exception:
            pass
    if old_qr_path and os.path.exists(old_qr_path):
        try:
            os.remove(old_qr_path)
        except Exception:
            pass
    if os.path.exists(old_word_path):
        try:
            os.remove(old_word_path)
        except Exception:
            pass

    # Clear fields
    cert.pdf_file = None
    cert.qr_code_image = None
    cert.save()

    # 2. Regenerate with new ID
    try:
        generate_pdf_and_qr(cert)
        cert.save()
    except Exception as e:
        return Response({"detail": f"Fayllarni generatsiya qilishda xatolik: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({
        "detail": "Sertifikat ID muvaffaqiyatli o'zgartirildi va hujjatlar qayta generatsiya qilindi.",
        "certificate_id": cert.certificate_id,
        "pdf_file": request.build_absolute_uri(cert.pdf_file.url) if cert.pdf_file else None,
        "qr_code_image": request.build_absolute_uri(cert.qr_code_image.url) if cert.qr_code_image else None,
    })



